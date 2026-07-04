import requests
from bs4 import BeautifulSoup

HEADERS = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}

# Parse SIH 2024 Excel
print("="*60)
print("SIH 2024 - Parsing Excel")
try:
    import openpyxl
    wb = openpyxl.load_workbook("data/SIH_PS_2024.xlsx")
    sheet = wb.active
    print(f"Rows: {sheet.max_row}, Cols: {sheet.max_column}")
    print(f"\nSheet name: {sheet.title}")
    print(f"\nFirst 5 rows:")
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i > 5:
            break
        print(f"  Row {i}: {[str(c)[:60] if c else '' for c in row[:10]]}")
    
    print(f"\nHeaders (Row 1):")
    for cell in sheet[1]:
        if cell.value:
            print(f"  {cell.column_letter}: {cell.value}")
    
    # Count actual data rows
    data_rows = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            data_rows += 1
    print(f"\nData rows: {data_rows}")
    
    # Sample a data row
    print(f"\nSample data row (row 2):")
    for cell in sheet[2]:
        if cell.value:
            print(f"  {cell.column_letter}: {str(cell.value)[:100]}")
            
except ImportError:
    print("openpyxl not installed, installing...")
    import subprocess
    subprocess.run(["python", "-m", "pip", "install", "openpyxl"])
    print("Run the script again after installation")

# Check SIH 2025 PS for simpler data format
print("\n" + "="*60)
print("SIH 2025 PS - Looking for patterns")
r = requests.get("https://www.sih.gov.in/sih2025PS", headers=HEADERS, timeout=15)
soup = BeautifulSoup(r.text, "lxml")

# Find all <th> with "Problem Statement ID" - each starts a new problem
th_psid = soup.find_all("th", string=lambda s: s and "Problem Statement ID" in s.strip())
print(f"Found {len(th_psid)} 'Problem Statement ID' headers")

if th_psid:
    # Each problem starts with a row containing "Problem Statement ID" <th>
    # Let's examine the first few
    for th in th_psid[:3]:
        row = th.find_parent("tr")
        print(f"\n--- PS Row ---")
        print(str(row)[:500])
        # Get all cells
        cells = row.find_all(["td", "th"])
        for cell in cells:
            print(f"  <{cell.name}> {cell.get_text(strip=True)[:80]}")

# Also try to find a pattern - maybe each problem is in its own <table>
print(f"\n--- Checking for per-problem tables ---")
subtables = soup.find_all("table")
print(f"Total tables: {len(subtables)}")
# Look at the table structure more carefully
# The main listing might have different structure
first_table = subtables[0] if subtables else None
if first_table:
    # Get all rows that have "Problem Statement ID" as first cell
    for tr in first_table.find_all("tr"):
        first_cell = tr.find(["th", "td"])
        if first_cell and "Problem Statement ID" in first_cell.get_text(strip=True):
            # This is a problem start row
            id_cell = tr.find("td")
            if id_cell:
                ps_id = id_cell.get_text(strip=True)
                # Get next rows for title and description
                next_tr = tr.find_next_sibling("tr")
                title = ""
                desc = ""
                org = ""
                category = ""
                theme = ""
                if next_tr:
                    th_title = next_tr.find("th")
                    if th_title and "Title" in th_title.get_text(strip=True):
                        title = next_tr.find("td").get_text(strip=True) if next_tr.find("td") else ""
                        next_tr2 = next_tr.find_next_sibling("tr")
                        if next_tr2:
                            th_desc = next_tr2.find("th")
                            if th_desc and "Description" in th_desc.get_text(strip=True):
                                desc_td = next_tr2.find("td")
                                if desc_td:
                                    desc = desc_td.get_text(strip=True)
                print(f"\n  PS: {ps_id}")
                print(f"  Title: {title[:80]}")
                print(f"  Desc: {desc[:100]}")
                break
